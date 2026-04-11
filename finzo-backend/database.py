

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import uuid
from datetime import datetime
from copy import copy



#  Column definitions


TXN_COLS = [
    "id", "user_id", "date", "description", "amount", "balance",
    "type", "category", "flagged", "triggered_rules",
    "filename", "upload_id", "uploaded_at",
]

UPLOAD_COLS = [
    "upload_id", "user_id", "filename", "count", "created_at",
]

RESOLVED_COLS = [
    "user_id", "txn_id", "resolved_at",
]

PROFILE_COLS = [
    "user_id", "display_name", "email", "phone",
    "monthly_budget", "currency", "updated_at",
]



#  Styling helpers


HEADER_FILL   = PatternFill("solid", fgColor="1A1D2E")
HEADER_FONT   = Font(name="Calibri", bold=True, color="ADC6FF", size=10)
DEBIT_FILL    = PatternFill("solid", fgColor="2A1A1A")
CREDIT_FILL   = PatternFill("solid", fgColor="1A2A1A")
FLAGGED_FILL  = PatternFill("solid", fgColor="2A1410")
ALT_FILL      = PatternFill("solid", fgColor="1C1E22")
NORMAL_FILL   = PatternFill("solid", fgColor="16181C")
BORDER_SIDE   = Side(style="thin", color="2A2D35")

def _cell_border():
    return Border(
        left=BORDER_SIDE, right=BORDER_SIDE,
        top=BORDER_SIDE,  bottom=BORDER_SIDE,
    )

def _style_header_row(ws, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill   = copy(HEADER_FILL)
        cell.font   = copy(HEADER_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _cell_border()
    ws.row_dimensions[1].height = 22


def _style_data_row(ws, row_idx, row_data, sheet_type="transactions"):
    """Apply row-level colour coding."""
    for c in range(1, len(row_data) + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.border = _cell_border()
        cell.alignment = Alignment(vertical="center")

    if sheet_type == "transactions":
        txn_type = row_data[6] if len(row_data) > 6 else ""
        flagged  = row_data[8] if len(row_data) > 8 else False
        if flagged:
            fill = copy(FLAGGED_FILL)
        elif txn_type == "credit":
            fill = copy(CREDIT_FILL)
        elif txn_type == "debit":
            fill = copy(DEBIT_FILL)
        else:
            fill = copy(ALT_FILL) if row_idx % 2 == 0 else copy(NORMAL_FILL)
        for c in range(1, len(row_data) + 1):
            ws.cell(row=row_idx, column=c).fill = fill
    else:
        fill = copy(ALT_FILL) if row_idx % 2 == 0 else copy(NORMAL_FILL)
        for c in range(1, len(row_data) + 1):
            ws.cell(row=row_idx, column=c).fill = fill


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


#  FinzoDB class


class FinzoDB:
    def __init__(self, path: str):
        self.path = path
        self._ensure_workbook()

    # ── internal

    def _ensure_workbook(self):
        """Create workbook with correct sheets if it doesn't exist."""
        try:
            wb = openpyxl.load_workbook(self.path)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # remove default sheet

        sheet_init = {
            "transactions":   (TXN_COLS,      [8, 18, 12, 40, 14, 14, 8, 14, 8, 30, 22, 12, 22]),
            "uploads":        (UPLOAD_COLS,    [14, 18, 30, 8, 22]),
            "fraud_resolved": (RESOLVED_COLS,  [18, 10, 22]),
            "profiles":       (PROFILE_COLS,   [18, 22, 28, 14, 14, 8, 22]),
        }
        changed = False
        for name, (cols, widths) in sheet_init.items():
            if name not in wb.sheetnames:
                ws = wb.create_sheet(name)
                ws.freeze_panes = "A2"
                ws.append(cols)
                _style_header_row(ws, len(cols))
                _set_col_widths(ws, widths)
                changed = True

        if changed:
            wb.save(self.path)

    def _load(self):
        return openpyxl.load_workbook(self.path)

    def _save(self, wb):
        wb.save(self.path)

    def _sheet_to_dicts(self, ws):
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h) for h in rows[0]]
        result = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            d = dict(zip(headers, row))
            result.append(d)
        return result

    def _coerce(self, d):
        """Convert types for JSON safety."""
        out = {}
        for k, v in d.items():
            if isinstance(v, bool):
                out[k] = v
            elif v is None:
                out[k] = None
            elif isinstance(v, (int, float)):
                out[k] = v
            else:
                s = str(v)
                if s.lower() == "true":
                    out[k] = True
                elif s.lower() == "false":
                    out[k] = False
                else:
                    try:
                        out[k] = int(s) if "." not in s else float(s)
                    except (ValueError, TypeError):
                        out[k] = s
        return out

    # ── transactions 

    def get_transactions(self, user_id: str) -> list:
        wb = self._load()
        ws = wb["transactions"]
        all_rows = self._sheet_to_dicts(ws)
        result = [self._coerce(r) for r in all_rows if str(r.get("user_id", "")) == user_id]
        return result

    def save_transactions(self, user_id: str, txns: list, filename: str) -> str:
        upload_id = str(uuid.uuid4())[:8]
        wb = self._load()
        ws = wb["transactions"]

        # avoid duplicates
        existing_ids = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == user_id:
                existing_ids.add(row[0])

        row_num = ws.max_row + 1
        for t in txns:
            if t["id"] in existing_ids:
                continue  # skip duplicates
            t["upload_id"] = upload_id
            row_data = [t.get(c) for c in TXN_COLS]
            ws.append(row_data)
            _style_data_row(ws, row_num, row_data, "transactions")
            row_num += 1

        self._save(wb)
        return upload_id

    def update_transaction(self, user_id: str, txn_id: int, updates: dict):
        wb = self._load()
        ws = wb["transactions"]
        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == str(txn_id) and str(row[1].value) == user_id:
                for key, val in updates.items():
                    if key in headers:
                        col_idx = headers.index(key) + 1
                        row[col_idx - 1].value = val
                break
        self._save(wb)

    def delete_transaction(self, user_id: str, txn_id: int):
        wb = self._load()
        ws = wb["transactions"]
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[0].value) == str(txn_id) and str(row[1].value) == user_id:
                ws.delete_rows(i)
                break
        self._save(wb)

    # ── fraud resolved 

    def get_resolved_alerts(self, user_id: str) -> list:
        wb = self._load()
        ws = wb["fraud_resolved"]
        rows = self._sheet_to_dicts(ws)
        return [int(r["txn_id"]) for r in rows if str(r.get("user_id", "")) == user_id]

    def resolve_alert(self, user_id: str, txn_id: int):
        if txn_id in self.get_resolved_alerts(user_id):
            return
        wb = self._load()
        ws = wb["fraud_resolved"]
        row_data = [user_id, txn_id, datetime.now().isoformat()]
        ws.append(row_data)
        _style_data_row(ws, ws.max_row, row_data, "other")
        self._save(wb)

    def unresolve_alert(self, user_id: str, txn_id: int):
        wb = self._load()
        ws = wb["fraud_resolved"]
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[1].value) == str(txn_id) and str(row[0].value) == user_id:
                ws.delete_rows(i)
                break
        self._save(wb)

    # ── uploads 

    def log_upload(self, user_id: str, filename: str, count: int, upload_id: str):
        wb = self._load()
        ws = wb["uploads"]
        row_data = [upload_id, user_id, filename, count, datetime.now().isoformat()]
        ws.append(row_data)
        _style_data_row(ws, ws.max_row, row_data, "other")
        self._save(wb)

    def get_upload_history(self, user_id: str) -> list:
        wb = self._load()
        ws = wb["uploads"]
        rows = self._sheet_to_dicts(ws)
        result = [self._coerce(r) for r in rows if str(r.get("user_id", "")) == user_id]
        return sorted(result, key=lambda x: str(x.get("created_at", "")), reverse=True)

    def delete_upload(self, user_id: str, upload_id: str):
        """Delete an upload record and all its transactions."""
        wb = self._load()

        # Remove from uploads sheet
        ws_up = wb["uploads"]
        rows_to_del = [
            i for i, row in enumerate(ws_up.iter_rows(min_row=2), start=2)
            if str(row[0].value) == upload_id and str(row[1].value) == user_id
        ]
        for i in reversed(rows_to_del):
            ws_up.delete_rows(i)

        # Remove transactions belonging to this upload
        ws_tx = wb["transactions"]
        tx_rows_to_del = [
            i for i, row in enumerate(ws_tx.iter_rows(min_row=2), start=2)
            if str(row[11].value) == upload_id and str(row[1].value) == user_id
        ]
        for i in reversed(tx_rows_to_del):
            ws_tx.delete_rows(i)

        self._save(wb)

    # ── profiles

    def get_profile(self, user_id: str) -> dict:
        wb = self._load()
        ws = wb["profiles"]
        rows = self._sheet_to_dicts(ws)
        for r in rows:
            if str(r.get("user_id", "")) == user_id:
                return self._coerce(r)
        return {"user_id": user_id}

    def save_profile(self, user_id: str, data: dict):
        wb = self._load()
        ws = wb["profiles"]
        rows = list(ws.iter_rows(min_row=2))
        headers = [cell.value for cell in ws[1]]
        data["user_id"] = user_id
        data["updated_at"] = datetime.now().isoformat()

        for row in rows:
            if str(row[0].value) == user_id:
                for key, val in data.items():
                    if key in headers:
                        row[headers.index(key)].value = val
                self._save(wb)
                return

        # Insert new profile
        row_data = [data.get(c) for c in PROFILE_COLS]
        ws.append(row_data)
        _style_data_row(ws, ws.max_row, row_data, "other")
        self._save(wb)
